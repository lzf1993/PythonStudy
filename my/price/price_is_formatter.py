import openpyxl
import re

def format_is_price(price_str: str) -> str:
    """
    将价格字符串中的IS价格四舍五入为整数
    """
    # 使用正则表达式查找IS价格
    is_pattern = r'IS;(\d+\.\d+)'
    match = re.search(is_pattern, price_str)
    
    if match:
        # 获取IS的价格
        is_price = float(match.group(1))
        # 四舍五入为整数
        rounded_price = round(is_price)
        # 替换原价格
        price_str = re.sub(is_pattern, f'IS;{rounded_price}', price_str)
    
    return price_str

def process_excel(input_path: str, output_path: str):
    """
    处理Excel文件，将IS价格四舍五入为整数
    """
    # 加载工作簿
    workbook = openpyxl.load_workbook(input_path)
    sheet = workbook.active
    
    # 查找Price列
    price_col = None
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == 'Price':
            price_col = col
            break
    
    if not price_col:
        print("未找到Price列")
        return
    
    # 处理每一行
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=price_col)
        if cell.value:
            # 格式化IS价格
            new_price = format_is_price(str(cell.value))
            if new_price != cell.value:
                print(f"行 {row}: IS价格已更新")
                print(f"原价格: {cell.value}")
                print(f"新价格: {new_price}")
                print("-" * 50)
            cell.value = new_price
    
    # 保存修改后的文件
    workbook.save(output_path)
    print(f"\n处理完成！文件已保存为: {output_path}")
