import openpyxl
import warnings
import re
import sys

warnings.filterwarnings('ignore')


#这些路径请使用来比较新旧文件价格变化的，旧路径是数据源，新路径是生成后的，不是将要生成的
# 将要生成的再项目根目录
oldpath = "/Users/lzf2/PycharmProjects/PythonStudy/my/price/olds.xlsx"
standard_re = re.compile(r'CN;\d*.\d*')


class Price:
    def __init__(self, name, old_price, new_price, replace=True, preciseMatch=True):
        self.name = name
        self.old_price = old_price
        self.new_price = new_price
        self.replace = replace
        self.preciseMatch = preciseMatch

    def __repr__(self):
        return str(self)

    def __str__(self):
        return "name=%s, old_price=%s, new_price=%s, replace=%s, preciseMatch=%s" % (
            self.name, self.old_price, self.new_price, self.replace, self.preciseMatch)



priceDic = {
    "1": [Price("CN", 1, 1, replace=False), Price("RU", -1, 29.00)],
    "3": [Price("CN", 3, 3, replace=False), Price("RU", -1, 49.00)],
    "6": [Price("CN", 6, 6, replace=False), Price("RU", -1, 99.00)],
    "8": [Price("CN", 8, 8, replace=False), Price("RU", -1, 129.00)],
    "12": [Price("CN", 12, 12, replace=False), Price("RU", -1, 199.00)],
    "18": [Price("CN", 18, 18, replace=False),  Price("RU", -1, 299.00)],
    "30": [Price("CN", 30, 30, replace=False), Price("RU", -1, 499.00)],
    "40": [Price("CN", 40, 40, replace=False),  Price("RU", -1, 599.00)],
    "50": [Price("CN", 50, 50, replace=False),  Price("RU", -1, 799.00)],
    "78": [Price("CN", 78, 78, replace=False),  Price("RU", -1, 1190.00)],
    "88": [Price("CN", 88, 88, replace=False),  Price("RU", -1, 1290.00)],
    "128": [Price("CN", 128, 128, replace=False), Price("RU", -1, 1990.00)],
    "148": [Price("CN", 148, 148, replace=False),  Price("RU", -1, 2190.00)],
    "158": [Price("CN", 158, 158, replace=False), Price("RU", -1, 2390.00)],
    "348": [Price("CN", 348, 348, replace=False), Price("RU", -1, 5490.00)],
    "388": [Price("CN", 388, 388, replace=False),  Price("RU", -1, 5890.00)],
    "588": [Price("CN", 588, 588, replace=False), Price("RU", -1, 8790.00)]
    }


# 获取标准价格，用于priceDic取key 默认标准价格为CN
def getStandardPrice(price: str):
    return str(int(float(standard_re.search(price).group().split(";")[1])))


# 读取一个单元格的值
def readOneCellStr(row: int, col: str, path: str):
    book = openpyxl.load_workbook(path)
    sheet = book.active
    return sheet["%s%s" % (col, row)].value


# 读取一行价格 这个是测试读取是否ok
def readOneLinePrice(sheet, row: int, col: str, path: str, new_price: bool = False):
    # 获取商品信息
    product = sheet["A%s" % row].value
    # 读取价格
    price = readOneCellStr(row, col, path)
    # 分离价格
    price_list = price.split(";")
    # 根据标准价格获取要替换的价格
    pd = priceDic.get(getStandardPrice(price))
    result = "%s %s = " % (row, product)
    for index, s in enumerate(price_list):
        for p in pd:
            if p.preciseMatch:
                # 非精准匹配
                if p.name == s:
                    result = "%s %s : %s" % (result, s, price_list[index + 1])
            else:
                if new_price:
                    pp = p.new_price
                else:
                    pp = p.old_price
                if index + index % 2 + 1 < len(price_list):
                    if str(pp) == str(price_list[index + index % 2 + 1]):
                        print('%s %s %s %s' % (
                            p.name, pp, price_list[index + index % 2], price_list[index + index % 2 + 1]))

                # print(pp)
                # print(price_list[index+1])
                # if index+1 < len(price_list) and pp == price_list[index+1]:
                #     result = "%s %s : %s" % (result, s, price_list[index])
                #     print(result)

    # print(result)


# 读价格 读取所有价格，如果需要
def readAllPrice(path: str):
    book = openpyxl.load_workbook(path)
    sheet = book.active
    for i in range(3, 50):
        # 读取i行 D列的内容
        readOneLinePrice(sheet, i, "D", path)


# 是否是精准匹配
def isPerfectMatch(name: str):
    for p in priceDic.get("3"):
        if p.name == name:
            return True
    return False


# 此方法用于验证替换前后是否ok
def compareOldAndNewFile(old_path: str, new_path: str):
    book = openpyxl.load_workbook(old_path)
    sheet = book.active
    for i in range(3, 50):
        product = sheet["A%s" % i].value
        print(product)
        old_p = readOneCellStr(i, "D", old_path)
        new_p = readOneCellStr(i, "D", new_path)
        old_list = re.findall(r'\w*;\d*.\d*;?', old_p)
        new_list = re.findall(r'\w*;\d*.\d*;?', new_p)
        for i in range(0, len(old_list)):
            if old_list[i] != new_list[i]:
                # if isPerfectMatch(str(old_list[i]).split(";")[0]):
                print('%s: %s -> %s' % (
                    str(old_list[i]).split(";")[0], str(old_list[i]).split(";")[1], str(new_list[i]).split(";")[1]))
        print()


# 写入价格
def writePrice(path: str, save_name: str):
    book = openpyxl.load_workbook(path)
    sheet = book.active
    for i in range(3, 50):
        location = "D%s" % i
        price = str(sheet[location].value)
        price_list = priceDic.get(getStandardPrice(price))
        if price_list:
            for p in price_list:
                # 如果该价格替换
                if p.replace:
                    if p.preciseMatch:
                        print('精确匹配')
                        # 正则要替换的值 如CN;3.00
                        rp = '%s;\d*.\d*' % p.name
                        # 新价格str 如CN;3.00
                        np = '%s;%.2f' % (p.name, float(p.new_price))
                        old_price = re.search(rp, price).group()
                        # 替换字符串中的旧价格
                        price = re.sub(rp, np, price)
                        new_price = re.search(rp, price).group()
                        print("旧价格是 %s" % old_price)
                        print("新价格是 %s" % new_price)
                        print(rp)
                        print(np)
                    else:
                        if p.old_price == -1:
                            print("模糊匹配必须设置旧价格")
                            return
                        print('模糊匹配')
                        old_rp = '%s' % p.old_price
                        old_rn = '%s' % p.new_price
                        # print(p.old_price)
                        # print(p.new_price)
                        price = re.sub(old_rp, old_rn, price)
        # print(price)
        sheet[location].value = price
        # price = price.replace(";", "")
        # print(price)
        # print(standard_re.search(price))
        # print(re.sub(r"CN;*?","CN;199999", price))
        # print(price)
    book.save(save_name)


if __name__ == "__main__":
    # 仅读取价格打开此方法
    # readAllPrice(newpath)
    
    # 生成的中间文件路径
    news_file = "/Users/lzf2/PycharmProjects/PythonStudy/my/price/news.xlsx"
    last_news_file = "/Users/lzf2/PycharmProjects/PythonStudy/my/price/last_news.xlsx"
    
    # 1. 首先执行华为价格处理
    writePrice(oldpath, news_file)
    print("\n华为价格处理完成！")
    
    # 2. 然后处理 IS 价格
    import price_is_formatter
    price_is_formatter.process_excel(news_file, last_news_file)
    
    # 3. 如果需要比较新旧价格，可以取消下面的注释
    # compareOldAndNewFile(oldpath, last_news_file)
