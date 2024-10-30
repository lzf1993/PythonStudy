import openpyxl

# 打开第一个 Excel 文件
workbook1 = openpyxl.load_workbook('file1.xlsx')
sheet1 = workbook1.active

# 打开第二个 Excel 文件
workbook2 = openpyxl.load_workbook('file2.xlsx')
sheet2 = workbook2.active

# 读取第一个 Excel 文件中的第一行数据
data_to_append = []
for cell in sheet1[1]:
    data_to_append.append(cell.value)

# 指定要将数据拼接的行数（这里假设是第二行）
target_row = 2

# 将数据拼接到第二个 Excel 文件的指定行
for index, value in enumerate(data_to_append, start=1):
    sheet2.cell(row=target_row, column=index, value=value)

# 保存更改
workbook2.save('file2.xlsx')
